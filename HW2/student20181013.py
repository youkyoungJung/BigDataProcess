 #!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

#total
row_id = 1
sum_v = 0.0

for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row=row_id, column=3).value * 0.3
		sum_v += ws.cell(row=row_id, column=4).value * 0.35
		sum_v += ws.cell(row=row_id, column=5).value * 0.34
		sum_v += ws.cell(row=row_id, column=6).value 
		ws.cell(row=row_id, column=7).value = sum_v
	row_id += 1
	
#grade

#scores=> put total
row_id = 1
scores = []

for row in ws:
	if row_id != 1:
		scores.append(ws.cell(row=row_id, column=7).value)
	row_id += 1

#rankings
rankings = [None]*len(scores)
for i in range(len(scores)):
	rankings[i] = len(scores) #초깃값 설정
	
	for j in range(len(scores)):
		if scores[j] < scores[i]:
			rankings[i] -= 1			
		
#for i in range(len(rankings)):
print(rankings)

row_id = 2;
for i in range(len(scores)):
	ws.cell(row=row_id, column=9).value = rankings[i]
	row_id += 1
	
#grade부여
print("0" + " A+등수범위 ",len(scores)*0.3*0.5)
print(len(scores)*0.3*0.5, " A0등수범위 ",len(scores)*0.3)
print(len(scores)*0.3, "B+등수범위 ",len(scores)*0.7*0.5)
print(len(scores)*0.7*0.5, " B0등수범위 ",len(scores)*0.7)
print(len(scores)*0.7, " C+등수범위 ",len(scores)*0.85)


row_id = 2
for i in range(len(scores)):
	if  rankings[i] <= len(scores)*0.15:	
		ws.cell(row=row_id, column=8).value = 'A+'
		
	elif rankings[i] <= len(scores)*0.3:
		ws.cell(row=row_id, column=8).value = 'A0'
		
	elif rankings[i] <= len(scores)*0.35:
		ws.cell(row=row_id, column=8).value = 'B+'
		
	elif rankings[i] <= len(scores)*0.7:
		ws.cell(row=row_id, column=8).value = 'B0'
		
	elif rankings[i] <= len(scores)*0.85:
		ws.cell(row=row_id, column=8).value = 'C+'

	else:
		ws.cell(row=row_id, column=8).value = 'C0'
	row_id += 1

wb.save("student.xlsx")

